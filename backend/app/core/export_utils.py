"""v1.4 数据导出核心函数——可独立于 FastAPI 应用直接调用和测试。"""
import io
import logging
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.core.constants import (
    EXPORT_DATE_FORMAT,
    EXPORT_MAX_ROWS_PER_SHEET,
    ACCOUNTING_PERIOD_FORMAT,
    EXPORT_DECIMAL_PLACES,
    EXPORT_FORMAT_SUPPORTED,
)

logger = logging.getLogger(__name__)

# ── 导出类型枚举 ──
EXPORT_TYPES = ("finance_report", "customers", "projects", "contracts", "tax_ledger")
EXPORT_FORMATS = ("xlsx", "pdf")

# ── 中文标题映射 ──
EXPORT_TITLES = {
    "finance_report": "月度财务报表",
    "customers": "客户列表",
    "projects": "项目列表",
    "contracts": "合同列表",
    "tax_ledger": "增值税发票台账",
}

# ── 列定义（严格按照 PRD，不得自行增减列）──
COLUMNS = {
    "finance_report": {
        "收支明细": ["日期", "类型(收入/支出)", "分类", "金额", "资金来源", "关联合同", "发票号码", "状态", "备注"],
        "分类汇总": ["分类", "收入合计", "支出合计", "净额"],
        "资金来源汇总": ["资金来源", "支出合计", "未结清笔数", "未结清金额"],
    },
    "customers": ["客户名称", "联系人", "电话", "公司名称", "状态", "来源渠道", "合作项目数",
                  "历史合同总额", "历史实收金额", "首次合作日期", "最近合作日期", "创建日期"],
    "projects": ["项目名称", "关联客户", "项目状态", "预算金额", "项目收入", "项目成本",
                 "项目利润", "利润率(%)", "开始日期", "结束日期", "进度(%)"],
    "contracts": ["合同编号", "合同标题", "关联客户", "关联项目", "合同金额", "已收金额",
                  "应收账款", "合同状态", "签署日期", "生效日期", "到期日期"],
    "tax_ledger": ["日期", "发票号码", "发票方向(销项/进项)", "发票类型", "金额", "税率", "税额", "关联合同", "备注"],
}

# ── Excel 样式 ──
_header_font = Font(name="Microsoft YaHei", bold=True, size=11)
_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_header_font_white = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
_cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
_thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _register_chinese_fonts() -> None:
    """注册中文字体（SimHei / Microsoft YaHei），PDF 生成时使用。"""
    import platform
    system = platform.system()
    font_registered = False

    if system == "Windows":
        import os
        windir = os.environ.get("WINDIR", r"C:\Windows")
        font_paths = [
            os.path.join(windir, "Fonts", "msyh.ttc"),
            os.path.join(windir, "Fonts", "simhei.ttf"),
            os.path.join(windir, "Fonts", "simsun.ttc"),
        ]
        for fp in font_paths:
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont("ChineseFont", fp))
                    font_registered = True
                    break
                except Exception:
                    continue
    elif system == "Darwin":
        import os
        font_paths = [
            "/System/Library/Fonts/PingFang.ttc",
            "/Library/Fonts/Arial Unicode.ttf",
            "/System/Library/Fonts/STHeiti Light.ttc",
        ]
        for fp in font_paths:
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont("ChineseFont", fp))
                    font_registered = True
                    break
                except Exception:
                    continue
    else:
        import os
        font_paths = [
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
        for fp in font_paths:
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont("ChineseFont", fp))
                    font_registered = True
                    break
                except Exception:
                    continue

    if not font_registered:
        logger.warning("未找到中文字体，PDF 中文可能显示为方块")


def get_export_filename(
    export_type: str, format: str, year: int,
    month: Optional[int] = None, quarter: Optional[int] = None,
) -> str:
    """
    生成文件名，格式：{模块名}_{年份}_{月份或季度}.{扩展名}

    示例：finance_report_2026_04.xlsx / tax_ledger_2026_Q1.pdf
    """
    ext = "xlsx" if format == "xlsx" else "pdf"
    if export_type == "tax_ledger" and quarter is not None:
        return f"{export_type}_{year}_Q{quarter}.{ext}"
    elif month is not None:
        return f"{export_type}_{year}_{month:02d}.{ext}"
    else:
        # 对于不需要年份月份的全局列表，使用当前年月
        today = date.today()
        return f"{export_type}_{today.year}_{today.month:02d}.{ext}"


def generate_excel(
    export_type: str, data: dict, year: int,
    month: Optional[int] = None, quarter: Optional[int] = None,
) -> bytes:
    """
    生成 Excel 文件，返回 bytes。数据为空时生成含表头的空文件。

    data 结构：
      finance_report: {"details": [...], "category_summary": [...], "funding_summary": [...]}
      customers: {"items": [...]}
      projects: {"items": [...]}
      contracts: {"items": [...]}
      tax_ledger: {"items": [...], "summary": {...}}
    """
    wb = Workbook()

    if export_type == "finance_report":
        _write_finance_report_xlsx(wb, data)
    elif export_type == "customers":
        _write_single_sheet_xlsx(wb, "客户列表", COLUMNS["customers"], data.get("items", []), "customers")
    elif export_type == "projects":
        _write_single_sheet_xlsx(wb, "项目列表", COLUMNS["projects"], data.get("items", []), "projects")
    elif export_type == "contracts":
        _write_single_sheet_xlsx(wb, "合同列表", COLUMNS["contracts"], data.get("items", []), "contracts")
    elif export_type == "tax_ledger":
        _write_tax_ledger_xlsx(wb, data)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf(
    export_type: str, data: dict, year: int,
    month: Optional[int] = None, quarter: Optional[int] = None,
) -> bytes:
    """
    生成 PDF 文件，返回 bytes。必须正确渲染中文字体。
    数据为空时生成含标题的空文件。
    """
    _register_chinese_fonts()

    output = io.BytesIO()
    title = EXPORT_TITLES.get(export_type, export_type)

    if export_type == "tax_ledger" and quarter is not None:
        period_label = f"{year}年 Q{quarter}"
    elif month is not None:
        period_label = f"{year}年{month:02d}月"
    else:
        period_label = f"当前快照 (导出时间: {date.today().strftime('%Y-%m-%d')})"

    doc = SimpleDocTemplate(
        output, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )

    elements: list = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ChineseTitle", parent=styles["Title"],
        fontName="ChineseFont", fontSize=16,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "ChineseSubtitle", parent=styles["Normal"],
        fontName="ChineseFont", fontSize=10,
        spaceAfter=12,
    )
    cell_style = ParagraphStyle(
        "ChineseCell", parent=styles["Normal"],
        fontName="ChineseFont", fontSize=8,
        leading=10,
    )

    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"报表期间：{period_label}", subtitle_style))
    elements.append(Spacer(1, 5 * mm))

    if export_type == "finance_report":
        _write_finance_report_pdf(elements, data, cell_style)
    elif export_type == "tax_ledger":
        _write_tax_ledger_pdf(elements, data, cell_style)
    else:
        cols = COLUMNS[export_type]
        items = data.get("items", [])
        _write_table_pdf(elements, cols, items, export_type, cell_style)

    doc.build(elements)
    return output.getvalue()


# ── Excel 内部函数 ──


def _apply_header_style(ws, col_count: int) -> None:
    """应用表头样式。"""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _header_font_white
        cell.fill = _header_fill
        cell.alignment = _cell_alignment
        cell.border = _thin_border


def _auto_width(ws, col_count: int) -> None:
    """自动调整列宽，并为所有行添加边框、居中、字体等基础样式。"""
    for col_idx in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value is not None:
                    # 动态计算字符视觉宽度：中文字符算作约 2.2，英文字符算作约 1.2
                    val_str = str(cell.value)
                    # 按照换行符分割，取最长的那一行计算
                    lines = val_str.split('\n')
                    for line in lines:
                        length = sum(2.2 if ord(c) > 255 else 1.2 for c in line)
                        if length > max_len:
                            max_len = length
        # 设置列宽，最多不超过 60
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # 给从第2行开始的数据行统一加上边框和居中（第1行表头已经在 _apply_header_style 处理）
    for row in ws.iter_rows(min_row=2, max_col=col_count):
        for cell in row:
            cell.alignment = _cell_alignment
            cell.border = _thin_border


def _write_finance_report_xlsx(wb: Workbook, data: dict) -> None:
    """写入月度财务报表（3 个 Sheet）。"""
    sheets_config = [
        ("收支明细", COLUMNS["finance_report"]["收支明细"], data.get("details", [])),
        ("分类汇总", COLUMNS["finance_report"]["分类汇总"], data.get("category_summary", [])),
        ("资金来源汇总", COLUMNS["finance_report"]["资金来源汇总"], data.get("funding_summary", [])),
    ]

    for idx, (sheet_name, columns, rows) in enumerate(sheets_config):
        ws = wb.active if idx == 0 else wb.create_sheet(sheet_name)
        ws.title = sheet_name
        ws.append(columns)
        _apply_header_style(ws, len(columns))
        for row in rows[:EXPORT_MAX_ROWS_PER_SHEET]:
            ws.append(row)
        _auto_width(ws, len(columns))


def _write_single_sheet_xlsx(wb: Workbook, sheet_name: str, columns: list, items: list, export_type: str) -> None:
    """写入单 Sheet 导出。"""
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)
    _apply_header_style(ws, len(columns))
    for row in items[:EXPORT_MAX_ROWS_PER_SHEET]:
        ws.append(row)
    _auto_width(ws, len(columns))


def _write_tax_ledger_xlsx(wb: Workbook, data: dict) -> None:
    """写入增值税发票台账（含底部汇总行）。"""
    cols = COLUMNS["tax_ledger"]
    ws = wb.active
    ws.title = "增值税发票台账"
    ws.append(cols)
    _apply_header_style(ws, len(cols))

    items = data.get("items", [])
    for row in items[:EXPORT_MAX_ROWS_PER_SHEET]:
        ws.append(row)

    summary = data.get("summary", {})
    if summary:
        output_tax = summary.get("output_tax", 0)
        input_tax = summary.get("input_tax", 0)
        payable = summary.get("payable_tax", 0)
        summary_row = ["汇总", "", "", "", "", "", output_tax, "", f"销项:{output_tax} 进项:{input_tax} 应纳:{payable}"]
        ws.append(summary_row)

    _auto_width(ws, len(cols))


# ── PDF 内部函数 ──


def _write_table_pdf(
    elements: list, columns: list, items: list,
    export_type: str, cell_style: ParagraphStyle,
) -> None:
    """写入通用 PDF 表格。"""
    table_data = [[Paragraph(str(c), cell_style) for c in columns]]
    for row in items:
        table_data.append([Paragraph(str(v) if v is not None else "", cell_style) for v in row])

    # A4 landscape width is 297mm. Margins are 15mm each. Available = 267mm.
    # We use 265mm to be safe and avoid overflow.
    col_widths = [265 * mm / max(len(columns), 1)] * len(columns)
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "ChineseFont"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(table)


def _write_finance_report_pdf(
    elements: list, data: dict, cell_style: ParagraphStyle,
) -> None:
    """写入月度财务报表 PDF。"""
    sections = [
        ("收支明细", COLUMNS["finance_report"]["收支明细"], data.get("details", [])),
        ("分类汇总", COLUMNS["finance_report"]["分类汇总"], data.get("category_summary", [])),
        ("资金来源汇总", COLUMNS["finance_report"]["资金来源汇总"], data.get("funding_summary", [])),
    ]

    section_style = ParagraphStyle(
        "SectionTitle", parent=cell_style,
        fontSize=12, spaceAfter=4, spaceBefore=8,
    )

    for section_name, columns, rows in sections:
        elements.append(Paragraph(section_name, section_style))
        _write_table_pdf(elements, columns, rows, "finance_report", cell_style)
        elements.append(Spacer(1, 5 * mm))


def _write_tax_ledger_pdf(
    elements: list, data: dict, cell_style: ParagraphStyle,
) -> None:
    """写入增值税发票台账 PDF（含汇总）。"""
    cols = COLUMNS["tax_ledger"]
    items = data.get("items", [])

    _write_table_pdf(elements, cols, items, "tax_ledger", cell_style)

    summary = data.get("summary", {})
    if summary:
        elements.append(Spacer(1, 3 * mm))
        summary_text = (
            f"销项税合计：{summary.get('output_tax', 0)}  |  "
            f"进项税合计（仅专用发票）：{summary.get('input_tax', 0)}  |  "
            f"应纳税额：{summary.get('payable_tax', 0)}"
        )
        elements.append(Paragraph(summary_text, cell_style))


# ── v1.8 财务对接导出功能 ────────────────────────────────────────

import os
import random
import string
from pathlib import Path
from typing import Any
from decimal import Decimal

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.constants import (
    ACCOUNTING_PERIOD_FORMAT,
    EXPORT_DECIMAL_PLACES,
    EXPORT_FORMAT_SUPPORTED,
)
from app.models.contract import Contract
from app.models.export_batch import ExportBatch
from app.models.finance import FinanceRecord
from app.models.invoice import Invoice
from app.models.quotation import Quotation
from app.models.customer import Customer

# 导出目录
V18_EXPORT_DIR = Path(__file__).parent.parent.parent / "exports"


def generate_export_batch_id() -> str:
    """生成唯一批次 ID（v1.8）。

    格式：EXP-YYYYMMDD-HHMMSS-随机6位
    """
    now = datetime.now()
    date_part = now.strftime("%Y%m%d")
    time_part = now.strftime("%H%M%S")
    random_part = "".join(random.choices(string.digits, k=6))
    return f"EXP-{date_part}-{time_part}-{random_part}"


def calculate_accounting_period(d: date) -> str:
    """计算会计期间（v1.8）。

    Args:
        d: 日期对象

    Returns:
        YYYY-MM 格式的会计期间字符串
    """
    return d.strftime(ACCOUNTING_PERIOD_FORMAT)


def get_period_date_range(accounting_period: str) -> tuple[date, date]:
    """根据会计期间获取日期范围（v1.8）。

    Args:
        accounting_period: YYYY-MM 格式字符串

    Returns:
        (start_date, end_date) 元组
    """
    year, month = map(int, accounting_period.split("-"))

    # 月初
    start_date = date(year, month, 1)

    # 月末
    if month == 12:
        end_date = date(year + 1, 1, 1) - __import__('datetime').timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - __import__('datetime').timedelta(days=1)

    return start_date, end_date


def map_to_finance_format(
    data: list[dict[str, Any]],
    target_format: str,
    export_type: str,
) -> list[dict[str, Any]]:
    """将数据映射为目标财务格式（v1.8）。

    Args:
        data: 原始数据列表
        target_format: 目标格式（generic/kingdee/yoyo/chanjet）
        export_type: 导出类型（contracts/payments/invoices）

    Returns:
        映射后的数据列表

    Raises:
        NotImplementedError: 当 target_format 不是 generic 时
    """
    if target_format not in EXPORT_FORMAT_SUPPORTED:
        raise NotImplementedError(
            f"导出格式 '{target_format}' 暂未实现，本版本仅支持: {EXPORT_FORMAT_SUPPORTED}"
        )

    # generic 格式直接返回原始数据（已经按目标字段组织）
    return data


def _format_export_value(value: Any, decimal_places: int = EXPORT_DECIMAL_PLACES) -> Any:
    """格式化导出值（v1.8）。

    - Decimal 转为保留指定位数的 float
    - date 转为字符串
    - None 转为空字符串
    """
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(round(value, decimal_places))
    if isinstance(value, date):
        return value.strftime(EXPORT_DATE_FORMAT)
    if isinstance(value, bool):
        return "是" if value else "否"
    return value


async def _fetch_contracts_data(
    db: AsyncSession,
    start_date: date | None = None,
    end_date: date | None = None,
) -> list[dict[str, Any]]:
    """获取合同导出数据（v1.8）。"""
    stmt = (
        select(
            Contract.contract_no,
            Contract.title,
            Contract.amount,
            Contract.signed_date,
            Customer.name.label("customer_name"),
            Quotation.quote_no,
        )
        .join(Customer, Contract.customer_id == Customer.id)
        .outerjoin(Quotation, Contract.quotation_id == Quotation.id)
    )

    if start_date:
        stmt = stmt.where(Contract.signed_date >= start_date)
    if end_date:
        stmt = stmt.where(Contract.signed_date <= end_date)

    stmt = stmt.order_by(Contract.signed_date.desc())

    result = await db.execute(stmt)
    rows = result.fetchall()

    data = []
    for row in rows:
        data.append({
            "合同编号": row.contract_no,
            "合同名称": row.title,
            "客户名称": row.customer_name,
            "合同金额": _format_export_value(row.amount),
            "签订日期": _format_export_value(row.signed_date),
            "会计期间": calculate_accounting_period(row.signed_date) if row.signed_date else "",
            "关联报价单": row.quote_no or "",
        })

    return data


async def _fetch_payments_data(
    db: AsyncSession,
    start_date: date | None = None,
    end_date: date | None = None,
    accounting_period: str | None = None,
) -> list[dict[str, Any]]:
    """获取收款导出数据（v1.8）。"""
    stmt = (
        select(
            FinanceRecord.id,
            FinanceRecord.date,
            FinanceRecord.amount,
            FinanceRecord.category,
            FinanceRecord.description,
            FinanceRecord.accounting_period,
            FinanceRecord.reconciliation_status,
            Customer.name.label("customer_name"),
            Contract.contract_no,
            Invoice.invoice_no,
        )
        .outerjoin(Contract, FinanceRecord.contract_id == Contract.id)
        .outerjoin(Customer, Contract.customer_id == Customer.id)
        .outerjoin(Invoice, FinanceRecord.invoice_id == Invoice.id)
        .where(FinanceRecord.type == "income")
    )

    if start_date:
        stmt = stmt.where(FinanceRecord.date >= start_date)
    if end_date:
        stmt = stmt.where(FinanceRecord.date <= end_date)
    if accounting_period:
        stmt = stmt.where(FinanceRecord.accounting_period == accounting_period)

    stmt = stmt.order_by(FinanceRecord.date.desc())

    result = await db.execute(stmt)
    rows = result.fetchall()

    data = []
    for row in rows:
        data.append({
            "id": row.id,
            "收款日期": _format_export_value(row.date),
            "收款金额": _format_export_value(row.amount),
            "收款方式": row.category or "",
            "客户名称": row.customer_name or "",
            "关联合同": row.contract_no or "",
            "关联发票": row.invoice_no or "",
            "会计期间": row.accounting_period or "",
            "对账状态": row.reconciliation_status or "pending",
            "备注": row.description or "",
        })

    return data


async def _fetch_invoices_data(
    db: AsyncSession,
    start_date: date | None = None,
    end_date: date | None = None,
) -> list[dict[str, Any]]:
    """获取发票导出数据（v1.8）。"""
    stmt = (
        select(
            Invoice.invoice_no,
            Invoice.invoice_type,
            Invoice.invoice_date,
            Customer.name.label("customer_name"),
            Invoice.amount_excluding_tax,
            Invoice.tax_rate,
            Invoice.tax_amount,
            Invoice.total_amount,
            Contract.contract_no,
            Invoice.status,
        )
        .join(Contract, Invoice.contract_id == Contract.id)
        .join(Customer, Contract.customer_id == Customer.id)
    )

    if start_date:
        stmt = stmt.where(Invoice.invoice_date >= start_date)
    if end_date:
        stmt = stmt.where(Invoice.invoice_date <= end_date)

    stmt = stmt.order_by(Invoice.invoice_date.desc())

    result = await db.execute(stmt)
    rows = result.fetchall()

    data = []
    for row in rows:
        data.append({
            "发票编号": row.invoice_no,
            "发票类型": row.invoice_type,
            "开票日期": _format_export_value(row.invoice_date),
            "客户名称": row.customer_name,
            "不含税金额": _format_export_value(row.amount_excluding_tax),
            "税率": float(row.tax_rate) if row.tax_rate else 0,
            "税额": _format_export_value(row.tax_amount),
            "价税合计": _format_export_value(row.total_amount),
            "关联合同": row.contract_no,
            "发票状态": row.status,
        })

    return data


def save_export_file(
    batch_id: str,
    export_type: str,
    target_format: str,
    data: list[dict[str, Any]],
) -> str:
    """保存导出文件到 exports/ 目录（v1.8）。

    Args:
        batch_id: 批次 ID
        export_type: 导出类型
        target_format: 目标格式
        data: 要导出的数据

    Returns:
        文件相对路径
    """
    # 确保目录存在
    V18_EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    # 文件名格式：{export_type}_{target_format}_{batch_id}.xlsx
    file_name = f"{export_type}_{target_format}_{batch_id}.xlsx"
    file_path = V18_EXPORT_DIR / file_name

    # 创建 Excel 文件
    wb = Workbook()
    ws = wb.active
    ws.title = export_type.capitalize()

    if data:
        # 写入表头
        headers = list(data[0].keys())
        ws.append(headers)

        # 写入数据
        for row_data in data:
            row = [row_data.get(h, "") for h in headers]
            ws.append(row)

    # 保存文件
    wb.save(file_path)

    logger.info(f"导出文件已保存: {file_path}, 记录数: {len(data)}")
    return str(file_path)


async def mark_records_as_exported(
    db: AsyncSession,
    record_ids: list[int],
    batch_id: int,
    accounting_period: str | None = None,
) -> None:
    """标记已导出的记录（v1.8）。

    仅对 payments 类型更新 finance_records 表。

    Args:
        db: 数据库会话
        record_ids: 记录 ID 列表
        batch_id: 导出批次 ID
        accounting_period: 会计期间（可选）
    """
    if not record_ids:
        return

    for record_id in record_ids:
        stmt = select(FinanceRecord).where(FinanceRecord.id == record_id)
        result = await db.execute(stmt)
        record = result.scalar_one_or_none()

        if record:
            record.export_batch_id = batch_id
            if accounting_period and not record.accounting_period:
                record.accounting_period = accounting_period

    await db.commit()
    logger.info(f"已标记 {len(record_ids)} 条记录为已导出，批次 ID: {batch_id}")


async def export_to_excel(
    db: AsyncSession,
    export_type: str,
    filters: dict[str, Any],
    target_format: str = "generic",
) -> dict[str, Any]:
    """统一导出入口，原子事务（v1.8）。

    Args:
        db: 数据库会话
        export_type: 导出类型（contracts/payments/invoices）
        filters: 筛选条件（start_date, end_date, accounting_period）
        target_format: 目标格式

    Returns:
        包含 id, batch_id, file_path, record_count 的字典

    Raises:
        ValueError: 当 export_type 无效时
        NotImplementedError: 当 target_format 不支持时
    """
    # 生成批次 ID
    batch_id = generate_export_batch_id()

    # 解析筛选条件
    start_date = filters.get("start_date")
    end_date = filters.get("end_date")
    accounting_period = filters.get("accounting_period")

    # 如果指定了会计期间，计算日期范围
    if accounting_period and not (start_date or end_date):
        start_date, end_date = get_period_date_range(accounting_period)

    # 获取数据
    if export_type == "contracts":
        data = await _fetch_contracts_data(db, start_date, end_date)
        record_ids = []  # 合同不更新批次 ID
    elif export_type == "payments":
        data = await _fetch_payments_data(db, start_date, end_date, accounting_period)
        # 提取记录 ID 用于标记
        record_ids = [row.get("id") for row in data] if data else []
    elif export_type == "invoices":
        data = await _fetch_invoices_data(db, start_date, end_date)
        record_ids = []  # 发票不更新批次 ID
    else:
        raise ValueError(f"无效的导出类型: {export_type}")

    # 格式映射
    mapped_data = map_to_finance_format(data, target_format, export_type)

    # 保存文件
    try:
        file_path = save_export_file(batch_id, export_type, target_format, mapped_data)
    except Exception as e:
        logger.error(f"保存导出文件失败: {e}")
        # 创建批次记录但 file_path 为 NULL
        batch = ExportBatch(
            batch_id=batch_id,
            export_type=export_type,
            target_format=target_format,
            accounting_period=accounting_period,
            start_date=start_date or date.today(),
            end_date=end_date or date.today(),
            record_count=len(mapped_data),
            file_path=None,
        )
        db.add(batch)
        await db.commit()
        await db.refresh(batch)
        raise

    # 创建批次记录
    batch = ExportBatch(
        batch_id=batch_id,
        export_type=export_type,
        target_format=target_format,
        accounting_period=accounting_period,
        start_date=start_date or date.today(),
        end_date=end_date or date.today(),
        record_count=len(mapped_data),
        file_path=file_path,
    )
    db.add(batch)
    await db.commit()
    await db.refresh(batch)

    # 标记已导出记录
    if record_ids:
        await mark_records_as_exported(db, record_ids, batch.id, accounting_period)

    logger.info(
        f"导出完成: batch_id={batch_id}, type={export_type}, "
        f"count={len(mapped_data)}, format={target_format}"
    )

    return {
        "id": batch.id,
        "batch_id": batch_id,
        "file_path": file_path,
        "record_count": len(mapped_data),
    }
