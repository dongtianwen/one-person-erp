import logging
from datetime import date
from io import BytesIO

from sqlalchemy.ext.asyncio import AsyncSession

from app.models.rd_expense import RdExpense
from app.core.constants import (
    DECIMAL_PLACES,
    EXPORT_DATE_FORMAT,
    RD_CATEGORY_LABELS,
    RD_SUB_CATEGORY_LABELS,
    RD_STATUS_LABELS,
)

logger = logging.getLogger(__name__)


def build_rd_export_rows(records: list[RdExpense]) -> list[dict]:
    rows = []
    for r in records:
        rows.append({
            "研发费用编号": r.rd_no,
            "费用发生日期": r.expense_date.strftime(EXPORT_DATE_FORMAT) if r.expense_date else "",
            "会计期间": r.accounting_period or "",
            "费用大类": RD_CATEGORY_LABELS.get(r.rd_category, r.rd_category),
            "费用子类": RD_SUB_CATEGORY_LABELS.get(r.rd_sub_category, r.rd_sub_category or ""),
            "金额（不含税）": float(r.amount or 0),
            "税额": float(r.tax_amount or 0),
            "价税合计": float(r.total_amount or 0),
            "供应商/收款方": r.vendor_name or "",
            "凭证号/发票号": r.invoice_no or "",
            "是否取得凭证": "是" if r.has_invoice else ("否" if r.has_invoice is False else ""),
            "凭证类型": r.invoice_type or "",
            "所属项目": r.project.name if r.project else "",
            "关联合同": r.contract.title if r.contract else "",
            "关联支出记录ID": str(r.finance_record_id or ""),
            "费用说明": r.description or "",
            "状态": RD_STATUS_LABELS.get(r.status, r.status),
            "备注": r.notes or "",
        })
    return rows


async def write_rd_expense_excel(
    db: AsyncSession,
    records: list[RdExpense],
    summary: dict | None = None,
) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "汇总表"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    if summary:
        ws_summary["A1"] = "研发费用台账汇总"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary.merge_cells("A1:D1")

        ws_summary["A3"] = "统计期间"
        ws_summary["B3"] = summary.get("period", "")
        ws_summary["A4"] = "费用合计（不含税）"
        ws_summary["B4"] = summary.get("total_amount", 0)
        ws_summary["A5"] = "税额合计"
        ws_summary["B5"] = summary.get("total_tax_amount", 0)
        ws_summary["A6"] = "价税合计"
        ws_summary["B6"] = summary.get("total_grand", 0)

        headers = ["费用大类", "金额（不含税）", "税额", "价税合计", "记录数"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_summary.cell(row=8, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, cat in enumerate(summary.get("categories", []), 9):
            ws_summary.cell(row=row_idx, column=1, value=cat.get("category_label", "")).border = thin_border
            ws_summary.cell(row=row_idx, column=2, value=cat.get("amount", 0)).border = thin_border
            ws_summary.cell(row=row_idx, column=3, value=cat.get("tax_amount", 0)).border = thin_border
            ws_summary.cell(row=row_idx, column=4, value=cat.get("total_amount", 0)).border = thin_border
            ws_summary.cell(row=row_idx, column=5, value=cat.get("count", 0)).border = thin_border

        col_widths = [18, 18, 12, 14, 10]
        for i, w in enumerate(col_widths, 1):
            ws_summary.column_dimensions[get_column_letter(i)].width = w

    ws_detail = wb.create_sheet(title="明细表")
    detail_headers = [
        "研发费用编号", "费用发生日期", "会计期间", "费用大类", "费用子类",
        "金额（不含税）", "税额", "价税合计", "供应商/收款方",
        "凭证号/发票号", "是否取得凭证", "凭证类型", "所属项目",
        "关联合同", "关联支出记录ID", "费用说明", "状态", "备注",
    ]
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    rows = build_rd_export_rows(records)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(detail_headers, 1):
            val = row_data.get(key, "")
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    detail_col_widths = [20, 14, 12, 16, 16, 16, 10, 14, 18, 18, 12, 12, 20, 24, 16, 30, 10, 30]
    for i, w in enumerate(detail_col_widths, 1):
        ws_detail.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
